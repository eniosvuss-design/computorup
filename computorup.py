"""
ComputoRUP v0.2 – app unica, nessun modulo esterno.
Avvio: streamlit run computorup.py
"""

import streamlit as st
import sqlite3, os, re, io, subprocess, tempfile, json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════════════
# CONFIGURAZIONE
# ════════════════════════════════════════════════════════════════════════════

DB_FILE = os.path.join(os.path.dirname(__file__), "computorup.db")
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
EXPORT_DIR = os.path.join(os.path.dirname(__file__), "exports")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

st.set_page_config(page_title="ComputoRUP", page_icon="🏗️", layout="wide")

# ════════════════════════════════════════════════════════════════════════════
# DATABASE
# ════════════════════════════════════════════════════════════════════════════

def db():
    c = sqlite3.connect(DB_FILE, check_same_thread=False)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA journal_mode=WAL")
    return c

def init_db():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS prezziario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codice TEXT, descrizione TEXT, um TEXT,
            prezzo REAL, mo_pct REAL, categoria TEXT, note TEXT
        );
        CREATE TABLE IF NOT EXISTS computo_meta (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titolo TEXT, committente TEXT, localita TEXT,
            rup TEXT, data TEXT, creato TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS computo_voci (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            computo_id INTEGER, area TEXT, categoria TEXT, tipo TEXT,
            codice TEXT, descrizione TEXT, um TEXT,
            qty REAL DEFAULT 0, prezzo REAL DEFAULT 0,
            importo REAL DEFAULT 0, mo_pct REAL DEFAULT 0,
            mo_euro REAL DEFAULT 0, note TEXT, modificato INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS preventivo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            computo_id INTEGER, ditta TEXT, filename TEXT,
            creato TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS preventivo_voci (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prev_id INTEGER, descrizione TEXT, um TEXT,
            qty REAL, prezzo REAL, importo REAL,
            mo_pct REAL DEFAULT 0, match_id INTEGER,
            scostamento REAL, note TEXT
        );
        CREATE TABLE IF NOT EXISTS sopralluogo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            computo_id INTEGER, filename TEXT,
            creato TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS sopralluogo_interventi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sopralluogo_id INTEGER, tempo_sec REAL,
            frame_path TEXT, descrizione TEXT,
            priorita TEXT DEFAULT 'media', importato INTEGER DEFAULT 0
        );
        """)

init_db()

# ════════════════════════════════════════════════════════════════════════════
# PARSER PREZZIARIO FVG (pdftotext -layout)
# ════════════════════════════════════════════════════════════════════════════

CATEGORY_MAP = {
    "01":"Opere provvisionali","02":"Demolizioni","03":"Scavi e rilevati",
    "04":"Fondazioni speciali","05":"Cls e strutture","06":"Murature",
    "07":"Coperture e impermeabilizzazioni","08":"Pavimenti e rivestimenti",
    "09":"Intonaci e tinteggiature","10":"Bonifiche",
    "11":"Scavi di fondazione","12":"Strutture in legno",
    "13":"Strutture in acciaio","14":"Impianti idraulici",
    "15":"Impianti elettrici","16":"Impianti termici",
    "17":"Tubazioni","18":"Ascensori","19":"Arredo urbano",
    "20":"Strade","21":"Verde pubblico",
    "22":"Lattoneria e coperture metalliche","23":"Scale",
    "25":"Restauro","26":"Consolidamento",
    "31":"Serramenti esterni","32":"Serramenti interni",
    "35":"Controsoffitti","36":"Vetrate",
    "50":"Impianti idrosanitari","51":"Impianti gas",
    "52":"Antincendio","53":"Impianti elettrici interni",
    "54":"HVAC","55":"Fotovoltaico",
    "70":"Parchi e giardini","99":"Sicurezza cantiere",
}

RE_CODE  = re.compile(r'^([0-9]{1,2}\.[0-9]\.[A-Z]{1,2}[0-9]{1,2}(?:\.[0-9]{1,2}(?:\.[A-Z]+)?)?)\s{2,}(.*)$')
RE_PRICE = re.compile(r'^\s{60,}(\S+)\s+(\d{1,3}[,\.]\d{2})\s+%\s+(\d{1,8}[,\.]\d{2})\s*$')
RE_SKIP  = re.compile(r'PREZZARIO|Prezzario|CODICE\s+DESC|U\.M\.\s+%|Pag\.\s+\d+|Edizione 202|Lavori\s+Pubblici')

def pf(s):
    s = str(s).strip().replace('\xa0','')
    if re.match(r'^\d{1,3}(\.\d{3})+(,\d+)?$', s):
        s = s.replace('.','').replace(',','.')
    else:
        s = s.replace(',','.')
    try: return float(s)
    except: return 0.0

def cat(code):
    m = re.match(r'^(\d+)\.', code)
    return CATEGORY_MAP.get(m.group(1).zfill(2) if m else '', 'Altro') if m else 'Altro'

def parse_prezziario(pdf_bytes):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes); tmp = f.name
    try:
        r = subprocess.run(["pdftotext","-layout",tmp,"-"],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=300)
        text = r.stdout
    except FileNotFoundError:
        os.unlink(tmp)
        return [], "❌ pdftotext non trovato. Installa poppler: brew install poppler"
    except subprocess.TimeoutExpired:
        os.unlink(tmp)
        return [], "❌ Timeout"
    finally:
        try: os.unlink(tmp)
        except: pass

    items, cur = [], None
    for line in text.split('\n'):
        if RE_SKIP.search(line): continue
        pm = RE_PRICE.match(line)
        if pm and cur:
            cur['um'] = pm.group(1)
            cur['mo_pct'] = pf(pm.group(2))
            cur['prezzo'] = pf(pm.group(3))
            cur['descrizione'] = re.sub(r'\s+',' ', cur['descrizione']).strip()[:600]
            items.append(cur); cur = None; continue
        cm = RE_CODE.match(line)
        if cm:
            cur = None
            cur = {'codice':cm.group(1),'descrizione':cm.group(2).strip(),
                   'um':'','prezzo':0.0,'mo_pct':0.0,
                   'categoria':cat(cm.group(1)),'note':''}
            continue
        if cur and line.startswith('   ') and line.strip():
            s = line.strip()
            if not re.match(r'^\d{1,4}\s*$',s) and len(s)>2 and len(cur['descrizione'])<700:
                cur['descrizione'] += ' '+s
    return items, f"✅ Estratte {len(items):,} voci"


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════

def fill(c): return PatternFill("solid",fgColor=c)
def font(bold=False,color="000000",size=10): return Font(bold=bold,color=color,size=size,name="Calibri")
def border():
    s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)

def export_excel(meta, voci, prev_voci=None, prezziario=None):
    wb = Workbook()

    # ── COPERTINA ──
    ws = wb.active; ws.title = "Copertina"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.merge_cells("A1:B1")
    ws["A1"].value = "COMPUTO METRICO ESTIMATIVO"
    ws["A1"].font  = Font(bold=True,size=18,color="FFFFFF",name="Calibri")
    ws["A1"].fill  = fill("1F3864")
    ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A2:B2")
    ws["A2"].value = meta.get("titolo","")
    ws["A2"].font  = Font(bold=True,size=13,color="1F3864",name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 28
    for r,(lbl,val) in enumerate([
        ("Committente",meta.get("committente","")),
        ("Località",meta.get("localita","")),
        ("RUP / Tecnico",meta.get("rup","")),
        ("Data",meta.get("data","")),
        ("Elaborato con","ComputoRUP v0.2"),
    ],4):
        ws.cell(r,1,lbl).font = font(bold=True)
        ws.cell(r,2,val)
        ws.row_dimensions[r].height = 20

    # ── COMPUTO ──
    ws2 = wb.create_sheet("Computo")
    ws2.freeze_panes = "A2"
    cols = ["N.","Area","Categoria","Tipo","Codice","Descrizione","U.M.",
            "Quantità","Prezzo €","Importo €","MO %","MO €","Note"]
    widths=[5,18,18,16,16,50,7,10,14,14,8,12,22]
    for i,(c,w) in enumerate(zip(cols,widths),1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        cell = ws2.cell(1,i,c)
        cell.fill = fill("1F3864"); cell.font = font(True,"FFFFFF")
        cell.alignment = Alignment(horizontal="center",wrap_text=True)
        cell.border = border()
    ws2.row_dimensions[1].height = 28

    row=2; cur_area=None; n=0
    for v in sorted(voci, key=lambda x:(x.get("area",""),x.get("categoria",""))):
        area = v.get("area","")
        if area != cur_area:
            ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=len(cols))
            c=ws2.cell(row,1,f"▶  {area.upper()}")
            c.fill=fill("2F5496"); c.font=font(True,"FFFFFF",11)
            c.alignment=Alignment(horizontal="left",vertical="center")
            ws2.row_dimensions[row].height=22; row+=1; cur_area=area
        n+=1
        data=[n,v.get("area",""),v.get("categoria",""),v.get("tipo",""),
              v.get("codice",""),v.get("descrizione",""),v.get("um",""),
              v.get("qty",0),v.get("prezzo",0),v.get("importo",0),
              v.get("mo_pct",0),v.get("mo_euro",0),v.get("note","")]
        bg = "F2F2F2" if n%2==0 else None
        for ci,val in enumerate(data,1):
            cell=ws2.cell(row,ci,val); cell.border=border()
            if bg: cell.fill=fill(bg)
            cell.alignment=Alignment(wrap_text=(ci==6),vertical="top")
            if ci in(9,10,12): cell.number_format='€ #,##0.00'; cell.alignment=Alignment(horizontal="right")
            if ci==8: cell.number_format='#,##0.000'; cell.alignment=Alignment(horizontal="right")
            if ci==11: cell.number_format='0.0"%"'; cell.alignment=Alignment(horizontal="right")
        ws2.row_dimensions[row].height=36; row+=1

    row+=1
    ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=9)
    ws2.cell(row,1,"TOTALE COMPLESSIVO").fill=fill("1F3864")
    ws2.cell(row,1).font=font(True,"FFFFFF",11)
    ws2.cell(row,1).alignment=Alignment(horizontal="right")
    tot=ws2.cell(row,10,sum(v.get("importo",0) for v in voci))
    tot.number_format='€ #,##0.00'; tot.font=font(True,"FFFFFF",11)
    tot.fill=fill("1F3864"); tot.alignment=Alignment(horizontal="right")

    # ── RIEPILOGO AREE ──
    ws3 = wb.create_sheet("Riepilogo aree")
    ws3.freeze_panes="A2"
    for i,(h,w) in enumerate(zip(["Area","Importo €"],[35,18]),1):
        ws3.column_dimensions[get_column_letter(i)].width=w
        c=ws3.cell(1,i,h); c.fill=fill("1F3864"); c.font=font(True,"FFFFFF")
        c.alignment=Alignment(horizontal="center"); c.border=border()
    from collections import defaultdict
    aree = defaultdict(float)
    for v in voci: aree[v.get("area","—")] += v.get("importo",0)
    r=2
    for area,imp in sorted(aree.items()):
        ws3.cell(r,1,area).border=border()
        c=ws3.cell(r,2,imp); c.number_format='€ #,##0.00'
        c.alignment=Alignment(horizontal="right"); c.border=border(); r+=1
    ws3.cell(r,1,"TOTALE").font=font(True); ws3.cell(r,1).fill=fill("E2EFDA")
    tc=ws3.cell(r,2,sum(aree.values())); tc.number_format='€ #,##0.00'
    tc.font=font(True); tc.fill=fill("E2EFDA"); tc.alignment=Alignment(horizontal="right")

    # ── RIEPILOGO CATEGORIE ──
    ws4 = wb.create_sheet("Riepilogo categorie")
    ws4.freeze_panes="A2"
    for i,(h,w) in enumerate(zip(["Categoria","Importo €"],[35,18]),1):
        ws4.column_dimensions[get_column_letter(i)].width=w
        c=ws4.cell(1,i,h); c.fill=fill("1F3864"); c.font=font(True,"FFFFFF")
        c.alignment=Alignment(horizontal="center"); c.border=border()
    cats = defaultdict(float)
    for v in voci: cats[v.get("categoria","—")] += v.get("importo",0)
    r=2
    for cat_,imp in sorted(cats.items()):
        ws4.cell(r,1,cat_).border=border()
        c=ws4.cell(r,2,imp); c.number_format='€ #,##0.00'
        c.alignment=Alignment(horizontal="right"); c.border=border(); r+=1
    ws4.cell(r,1,"TOTALE").font=font(True); ws4.cell(r,1).fill=fill("E2EFDA")
    tc=ws4.cell(r,2,sum(cats.values())); tc.number_format='€ #,##0.00'
    tc.font=font(True); tc.fill=fill("E2EFDA"); tc.alignment=Alignment(horizontal="right")

    # ── SOGGETTO A RIBASSO ──
    for tipo,sheet_name in [("Soggetto a ribasso","Soggetto a ribasso"),
                             ("Non soggetto a ribasso","Non soggetto a ribasso")]:
        wsX = wb.create_sheet(sheet_name); wsX.freeze_panes="A2"
        for i,(h,w) in enumerate(zip(["N.","Codice","Descrizione","U.M.","Qty","Prezzo €","Importo €"],
                                      [5,16,50,7,10,14,14]),1):
            wsX.column_dimensions[get_column_letter(i)].width=w
            c=wsX.cell(1,i,h); c.fill=fill("1F3864"); c.font=font(True,"FFFFFF")
            c.alignment=Alignment(horizontal="center"); c.border=border()
        filtered=[v for v in voci if v.get("tipo","")==tipo]
        rr=2; tot_r=0
        for nn,v in enumerate(filtered,1):
            data=[nn,v.get("codice",""),v.get("descrizione",""),v.get("um",""),
                  v.get("qty",0),v.get("prezzo",0),v.get("importo",0)]
            for ci,val in enumerate(data,1):
                cell=wsX.cell(rr,ci,val); cell.border=border()
                cell.alignment=Alignment(wrap_text=(ci==3),vertical="top")
                if ci in(6,7): cell.number_format='€ #,##0.00'; cell.alignment=Alignment(horizontal="right")
            wsX.row_dimensions[rr].height=36; tot_r+=v.get("importo",0); rr+=1
        wsX.cell(rr,6,"TOTALE").font=font(True)
        tc=wsX.cell(rr,7,tot_r); tc.number_format='€ #,##0.00'
        tc.font=font(True); tc.fill=fill("E2EFDA"); tc.alignment=Alignment(horizontal="right")

    # ── SICUREZZA / MANODOPERA ──
    for cat_filter,sheet_name in [("Sicurezza cantiere","Sicurezza"),("Manodopera","Manodopera")]:
        wsX=wb.create_sheet(sheet_name); wsX.freeze_panes="A2"
        for i,(h,w) in enumerate(zip(["Codice","Descrizione","Importo €"],[16,55,14]),1):
            wsX.column_dimensions[get_column_letter(i)].width=w
            c=wsX.cell(1,i,h); c.fill=fill("1F3864"); c.font=font(True,"FFFFFF")
            c.alignment=Alignment(horizontal="center"); c.border=border()
        rr=2; tot_r=0
        for v in voci:
            if v.get("categoria","")==cat_filter:
                wsX.cell(rr,1,v.get("codice","")).border=border()
                wsX.cell(rr,2,v.get("descrizione","")).border=border()
                wsX.cell(rr,2).alignment=Alignment(wrap_text=True)
                c=wsX.cell(rr,3,v.get("importo",0)); c.number_format='€ #,##0.00'
                c.alignment=Alignment(horizontal="right"); c.border=border()
                tot_r+=v.get("importo",0); rr+=1
        if rr==2: wsX.cell(2,2,"Nessuna voce in questa categoria.")
        else:
            tc=wsX.cell(rr,3,tot_r); tc.number_format='€ #,##0.00'
            tc.font=font(True); tc.fill=fill("E2EFDA"); tc.alignment=Alignment(horizontal="right")

    # ── PREZZIARIO ──
    if prezziario:
        wsp=wb.create_sheet("Prezziario"); wsp.freeze_panes="A2"
        for i,(h,w) in enumerate(zip(["Codice","Descrizione","U.M.","Prezzo €","MO %","Categoria"],
                                      [16,55,7,14,8,25]),1):
            wsp.column_dimensions[get_column_letter(i)].width=w
            c=wsp.cell(1,i,h); c.fill=fill("1F3864"); c.font=font(True,"FFFFFF")
            c.alignment=Alignment(horizontal="center"); c.border=border()
        for rr,p in enumerate(prezziario,2):
            data=[p.get("codice",""),p.get("descrizione",""),p.get("um",""),
                  p.get("prezzo",0),p.get("mo_pct",0),p.get("categoria","")]
            for ci,val in enumerate(data,1):
                cell=wsp.cell(rr,ci,val); cell.border=border()
                cell.alignment=Alignment(wrap_text=(ci==2),vertical="top")
                if ci==4: cell.number_format='€ #,##0.00'; cell.alignment=Alignment(horizontal="right")
                if ci==5: cell.number_format='0.0"%"'; cell.alignment=Alignment(horizontal="right")
            wsp.row_dimensions[rr].height=28
        wsp.auto_filter.ref=f"A1:F{len(prezziario)+1}"

    # ── CONFRONTO PREVENTIVO ──
    if prev_voci:
        wsc=wb.create_sheet("Confronto preventivo"); wsc.freeze_panes="A2"
        heads=["Descrizione preventivo","U.M.","Qty","Prezzo prev €",
               "Importo prev €","Codice prezziario","Prezzo prezz €","Diff €","Scost %","Note"]
        widths2=[45,7,10,14,16,16,14,12,10,25]
        for i,(h,w) in enumerate(zip(heads,widths2),1):
            wsc.column_dimensions[get_column_letter(i)].width=w
            c=wsc.cell(1,i,h); c.fill=fill("1F3864"); c.font=font(True,"FFFFFF")
            c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=border()
        for rr,v in enumerate(prev_voci,2):
            sc=v.get("scostamento")
            data=[v.get("descrizione",""),v.get("um",""),v.get("qty"),
                  v.get("prezzo",0),v.get("importo",0),
                  v.get("codice_match",""),v.get("prezzo_match",0),
                  v.get("diff"),
                  sc/100 if sc else None,
                  v.get("note","")]
            for ci,val in enumerate(data,1):
                cell=wsc.cell(rr,ci,val); cell.border=border()
                cell.alignment=Alignment(wrap_text=(ci==1),vertical="top")
                if ci in(4,5,7,8): cell.number_format='€ #,##0.00'; cell.alignment=Alignment(horizontal="right")
                if ci==9 and val is not None:
                    cell.number_format='0.0%'; cell.alignment=Alignment(horizontal="right")
                    if sc and sc>20: cell.fill=fill("FCE4D6")
            wsc.row_dimensions[rr].height=34

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# VIDEO – ESTRAZIONE FRAME E INTERVENTI
# ════════════════════════════════════════════════════════════════════════════

def extract_frames(video_bytes, interval_sec=5, max_frames=60):
    """
    Estrae frame dal video ogni interval_sec secondi.
    Ritorna lista di (tempo_sec, jpeg_bytes).
    """
    try:
        import cv2, numpy as np
    except ImportError:
        return None, "opencv non installato. Esegui: pip install opencv-python-headless"

    with tempfile.NamedTemporaryFile(suffix=".mp4", delete=False) as f:
        f.write(video_bytes); tmp = f.name
    try:
        cap = cv2.VideoCapture(tmp)
        fps = cap.get(cv2.CAP_PROP_FPS) or 25
        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = total / fps
        frames = []
        step = int(fps * interval_sec)
        frame_idx = 0
        while len(frames) < max_frames:
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
            ret, frm = cap.read()
            if not ret: break
            # resize per leggerezza
            h, w = frm.shape[:2]
            scale = min(640/w, 360/h, 1.0)
            frm = cv2.resize(frm, (int(w*scale), int(h*scale)))
            _, buf = cv2.imencode(".jpg", frm, [cv2.IMWRITE_JPEG_QUALITY, 80])
            t = frame_idx / fps
            frames.append((t, buf.tobytes()))
            frame_idx += step
        cap.release()
        return frames, f"✅ Estratti {len(frames)} frame (durata video: {duration:.0f}s)"
    finally:
        os.unlink(tmp)


# ════════════════════════════════════════════════════════════════════════════
# RICERCA FUZZY
# ════════════════════════════════════════════════════════════════════════════

def fuzzy_search(items, query, limit=8):
    if not query or not items: return items[:limit]
    try:
        from rapidfuzz import process, fuzz
        descs = [i.get("descrizione","") for i in items]
        hits  = process.extract(query, descs, scorer=fuzz.partial_ratio,
                                limit=limit, score_cutoff=30)
        return [items[h[2]] for h in hits]
    except ImportError:
        kw = query.lower().split()
        scored = [(sum(1 for k in kw if k in i.get("descrizione","").lower()), i) for i in items]
        return [i for s,i in sorted(scored,key=lambda x:-x[0]) if s>0][:limit]


# ════════════════════════════════════════════════════════════════════════════
# NAVIGAZIONE
# ════════════════════════════════════════════════════════════════════════════

PAGES = [
    "🏠 Home",
    "📚 Prezziario",
    "📋 Computo",
    "🔍 Cerca & Aggiungi",
    "🧾 Preventivo",
    "⚖️ Confronto",
    "🎬 Sopralluogo Video",
    "📤 Esporta",
]

if "page" not in st.session_state:
    st.session_state.page = "🏠 Home"

with st.sidebar:
    st.title("🏗️ ComputoRUP")
    st.caption("v0.2 – uso locale")
    st.divider()
    for p in PAGES:
        if st.button(p, use_container_width=True,
                     type="primary" if st.session_state.page==p else "secondary"):
            st.session_state.page = p
            st.rerun()
    st.divider()
    # selezione computo attivo
    with db() as conn:
        computi = [dict(r) for r in conn.execute("SELECT * FROM computo_meta ORDER BY creato DESC")]
    if computi:
        opts = {f"[{c['id']}] {c['titolo']}": c["id"] for c in computi}
        label = st.selectbox("Computo attivo", list(opts.keys()),
                             key="computo_sel_sidebar")
        st.session_state["computo_id"] = opts[label]
    else:
        st.info("Nessun computo. Creane uno.")
        st.session_state["computo_id"] = None

page = st.session_state.page


# ════════════════════════════════════════════════════════════════════════════
# 🏠 HOME
# ════════════════════════════════════════════════════════════════════════════

if page == "🏠 Home":
    st.title("🏗️ ComputoRUP v0.2")
    st.info("Strumento locale per computi metrici estimativi – Prezzario FVG 2025")

    with db() as conn:
        n_voci_prez = conn.execute("SELECT COUNT(*) FROM prezziario").fetchone()[0]
        n_computi   = conn.execute("SELECT COUNT(*) FROM computo_meta").fetchone()[0]
        n_voci_comp = conn.execute("SELECT COUNT(*) FROM computo_voci").fetchone()[0]
        tot_imp     = conn.execute("SELECT COALESCE(SUM(importo),0) FROM computo_voci").fetchone()[0]

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Voci prezziario",  f"{n_voci_prez:,}")
    c2.metric("Computi aperti",   n_computi)
    c3.metric("Voci nei computi", n_voci_comp)
    c4.metric("Importo totale",   f"€ {tot_imp:,.2f}")

    st.subheader("Flusso di lavoro")
    st.markdown("""
    1. **📚 Prezziario** → carica il PDF del Prezzario FVG (6.500 voci in ~6 sec)
    2. **📋 Computo** → crea il computo e inserisci l'anagrafica
    3. **🔍 Cerca & Aggiungi** → cerca le voci e aggiungile al computo
    4. **🎬 Sopralluogo Video** → carica un video di sopralluogo, segna gli interventi frame per frame
    5. **🧾 Preventivo** → carica il preventivo della ditta
    6. **⚖️ Confronto** → abbina le voci e calcola gli scostamenti
    7. **📤 Esporta** → genera l'Excel professionale
    """)


# ════════════════════════════════════════════════════════════════════════════
# 📚 PREZZIARIO
# ════════════════════════════════════════════════════════════════════════════

elif page == "📚 Prezziario":
    st.title("📚 Prezziario FVG")

    with db() as conn:
        n = conn.execute("SELECT COUNT(*) FROM prezziario").fetchone()[0]

    if n > 0:
        st.success(f"Prezziario caricato: **{n:,} voci**")
        if st.button("🗑️ Cancella e ricarica"):
            with db() as conn:
                conn.execute("DELETE FROM prezziario"); conn.commit()
            st.rerun()
    else:
        st.info("Nessun prezziario. Carica il PDF del Prezzario FVG 2025.")
        uploaded = st.file_uploader("PDF Prezzario FVG", type=["pdf"])
        if uploaded and st.button("⚙️ Estrai voci", type="primary"):
            with st.spinner("Lettura PDF… (circa 6-10 secondi)"):
                items, msg = parse_prezziario(uploaded.read())
            if items:
                st.success(msg)
                with db() as conn:
                    conn.execute("DELETE FROM prezziario")
                    conn.executemany(
                        "INSERT INTO prezziario(codice,descrizione,um,prezzo,mo_pct,categoria,note) VALUES(?,?,?,?,?,?,?)",
                        [(i["codice"],i["descrizione"],i["um"],i["prezzo"],
                          i["mo_pct"],i["categoria"],i.get("note","")) for i in items]
                    )
                    conn.commit()
                st.rerun()
            else:
                st.error(msg)

    if n > 0:
        st.subheader("Cerca nel prezziario")
        q = st.text_input("Codice o parola chiave", placeholder="es: guaina bituminosa")
        cat_f = st.selectbox("Categoria", ["(tutte)"] + list(CATEGORY_MAP.values()))
        with db() as conn:
            if q:
                rows = [dict(r) for r in conn.execute(
                    "SELECT * FROM prezziario WHERE codice LIKE ? OR descrizione LIKE ? LIMIT 100",
                    (f"%{q}%", f"%{q}%"))]
            else:
                if cat_f != "(tutte)":
                    rows = [dict(r) for r in conn.execute(
                        "SELECT * FROM prezziario WHERE categoria=? LIMIT 200", (cat_f,))]
                else:
                    rows = [dict(r) for r in conn.execute("SELECT * FROM prezziario LIMIT 100")]

        if rows:
            df = pd.DataFrame(rows)[["id","codice","descrizione","um","prezzo","mo_pct","categoria"]]
            df.columns = ["ID","Codice","Descrizione","U.M.","Prezzo €","MO %","Categoria"]
            st.dataframe(df, use_container_width=True, height=400,
                         column_config={
                             "Prezzo €": st.column_config.NumberColumn(format="€ %.2f"),
                             "MO %": st.column_config.NumberColumn(format="%.1f%%"),
                         })
        else:
            st.info("Nessun risultato.")

        # Modifica manuale
        with st.expander("✏️ Modifica / aggiungi voce manuale"):
            with st.form("add_prez"):
                c1,c2 = st.columns(2)
                nc = c1.text_input("Codice")
                nu = c2.text_input("U.M.")
                nd = st.text_area("Descrizione")
                c3,c4,c5 = st.columns(3)
                np_ = c3.number_input("Prezzo €", min_value=0.0, format="%.2f")
                nmo = c4.number_input("Incid. MO %", min_value=0.0, max_value=100.0, format="%.1f")
                ncat= c5.text_input("Categoria")
                if st.form_submit_button("Aggiungi"):
                    with db() as conn:
                        conn.execute(
                            "INSERT INTO prezziario(codice,descrizione,um,prezzo,mo_pct,categoria,note) VALUES(?,?,?,?,?,?,?)",
                            (nc,nd,nu,np_,nmo,ncat,""))
                        conn.commit()
                    st.success("Voce aggiunta."); st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# 📋 COMPUTO
# ════════════════════════════════════════════════════════════════════════════

elif page == "📋 Computo":
    st.title("📋 Computo Metrico Estimativo")

    # Anagrafica
    st.subheader("Anagrafica intervento")
    cid = st.session_state.get("computo_id")
    meta = {}
    if cid:
        with db() as conn:
            row = conn.execute("SELECT * FROM computo_meta WHERE id=?", (cid,)).fetchone()
            if row: meta = dict(row)

    with st.form("meta_form"):
        c1,c2 = st.columns(2)
        titolo = c1.text_input("Titolo intervento *", value=meta.get("titolo",""))
        comm   = c2.text_input("Committente", value=meta.get("committente",""))
        loc    = c1.text_input("Località", value=meta.get("localita",""))
        rup    = c2.text_input("RUP / Tecnico", value=meta.get("rup",""))
        data_  = c1.text_input("Data", value=meta.get("data", datetime.today().strftime("%d/%m/%Y")))
        if st.form_submit_button("💾 Salva anagrafica", type="primary"):
            if not titolo:
                st.error("Il titolo è obbligatorio.")
            else:
                with db() as conn:
                    if cid:
                        conn.execute("UPDATE computo_meta SET titolo=?,committente=?,localita=?,rup=?,data=? WHERE id=?",
                                     (titolo,comm,loc,rup,data_,cid))
                    else:
                        conn.execute("INSERT INTO computo_meta(titolo,committente,localita,rup,data) VALUES(?,?,?,?,?)",
                                     (titolo,comm,loc,rup,data_))
                        cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    conn.commit()
                st.session_state["computo_id"] = cid
                st.success(f"Salvato (ID: {cid})."); st.rerun()

    if not cid:
        st.info("Salva prima l'anagrafica.")
        st.stop()

    # Aggiungi voce manuale
    st.subheader("Aggiungi voce")
    with st.expander("➕ Nuova voce manuale"):
        with st.form("add_voce"):
            c1,c2,c3 = st.columns(3)
            area = c1.text_input("Area intervento", value="Zona A")
            cat_ = c2.selectbox("Categoria",["OG1","OG2","OG11","Sicurezza cantiere","Manodopera","Altro"])
            tipo = c3.selectbox("Tipo",["Soggetto a ribasso","Non soggetto a ribasso"])
            c4,c5 = st.columns(2)
            cod  = c4.text_input("Codice prezziario")
            um   = c5.text_input("U.M.")
            desc = st.text_area("Descrizione *")
            c6,c7,c8 = st.columns(3)
            qty  = c6.number_input("Quantità", min_value=0.0, value=1.0, format="%.3f")
            prz  = c7.number_input("Prezzo €", min_value=0.0, format="%.2f")
            mo   = c8.number_input("MO %", min_value=0.0, max_value=100.0, format="%.1f")
            note = st.text_input("Note RUP")
            st.metric("Importo", f"€ {qty*prz:,.2f}")
            if st.form_submit_button("Aggiungi", type="primary"):
                if not desc:
                    st.error("Descrizione obbligatoria.")
                else:
                    imp  = qty*prz; mo_e = imp*mo/100
                    with db() as conn:
                        conn.execute(
                            "INSERT INTO computo_voci(computo_id,area,categoria,tipo,codice,descrizione,um,qty,prezzo,importo,mo_pct,mo_euro,note) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (cid,area,cat_,tipo,cod,desc,um,qty,prz,imp,mo,mo_e,note))
                        conn.commit()
                    st.success("Voce aggiunta."); st.rerun()

    # Tabella voci
    st.subheader("Voci del computo")
    with db() as conn:
        voci = [dict(r) for r in conn.execute(
            "SELECT * FROM computo_voci WHERE computo_id=? ORDER BY area,categoria,id", (cid,))]

    if not voci:
        st.info("Nessuna voce. Aggiungile qui sopra o con Cerca & Aggiungi.")
    else:
        tot   = sum(v["importo"] for v in voci)
        mo_t  = sum(v["mo_euro"] for v in voci)
        rib   = sum(v["importo"] for v in voci if v["tipo"]=="Soggetto a ribasso")
        nrib  = sum(v["importo"] for v in voci if v["tipo"]=="Non soggetto a ribasso")
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Totale",f"€ {tot:,.2f}")
        c2.metric("Manodopera",f"€ {mo_t:,.2f}")
        c3.metric("A ribasso",f"€ {rib:,.2f}")
        c4.metric("Non a ribasso",f"€ {nrib:,.2f}")

        df = pd.DataFrame(voci)
        edited = st.data_editor(
            df[["id","area","categoria","tipo","codice","descrizione","um","qty","prezzo","importo","mo_pct","mo_euro","note"]],
            use_container_width=True, num_rows="dynamic", disabled=["id","importo","mo_euro"],
            column_config={
                "id":       st.column_config.NumberColumn("ID",width="small"),
                "area":     st.column_config.TextColumn("Area"),
                "categoria":st.column_config.SelectboxColumn("Categoria",options=["OG1","OG2","OG11","Sicurezza cantiere","Manodopera","Altro"]),
                "tipo":     st.column_config.SelectboxColumn("Tipo",options=["Soggetto a ribasso","Non soggetto a ribasso"]),
                "codice":   st.column_config.TextColumn("Codice"),
                "descrizione":st.column_config.TextColumn("Descrizione",width="large"),
                "um":       st.column_config.TextColumn("U.M.",width="small"),
                "qty":      st.column_config.NumberColumn("Qty",format="%.3f"),
                "prezzo":   st.column_config.NumberColumn("Prezzo €",format="%.2f"),
                "importo":  st.column_config.NumberColumn("Importo €",format="%.2f"),
                "mo_pct":   st.column_config.NumberColumn("MO %",format="%.1f"),
                "mo_euro":  st.column_config.NumberColumn("MO €",format="%.2f"),
                "note":     st.column_config.TextColumn("Note"),
            }, height=450,
        )
        c_s, c_d = st.columns([3,1])
        if c_s.button("💾 Salva modifiche", type="primary"):
            with db() as conn:
                for row in edited.to_dict(orient="records"):
                    if pd.notna(row.get("id")) and row.get("id"):
                        q2=float(row.get("qty",0) or 0); p2=float(row.get("prezzo",0) or 0)
                        m2=float(row.get("mo_pct",0) or 0); imp2=q2*p2; mo2=imp2*m2/100
                        conn.execute(
                            "UPDATE computo_voci SET area=?,categoria=?,tipo=?,codice=?,descrizione=?,um=?,qty=?,prezzo=?,importo=?,mo_pct=?,mo_euro=?,note=? WHERE id=?",
                            (row.get("area",""),row.get("categoria",""),row.get("tipo",""),
                             row.get("codice",""),row.get("descrizione",""),row.get("um",""),
                             q2,p2,imp2,m2,mo2,row.get("note",""),int(row["id"])))
                conn.commit()
            st.success("Salvato."); st.rerun()
        with c_d.expander("🗑️ Elimina per ID"):
            del_id = st.number_input("ID",min_value=1,step=1,key="del_id")
            if st.button("Elimina"):
                with db() as conn:
                    conn.execute("DELETE FROM computo_voci WHERE id=?",(int(del_id),)); conn.commit()
                st.success(f"Eliminata voce {del_id}."); st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# 🔍 CERCA & AGGIUNGI
# ════════════════════════════════════════════════════════════════════════════

elif page == "🔍 Cerca & Aggiungi":
    st.title("🔍 Cerca Prezziario e Aggiungi al Computo")
    cid = st.session_state.get("computo_id")
    if not cid:
        st.warning("Seleziona un computo dalla barra laterale."); st.stop()

    with db() as conn:
        n_prez = conn.execute("SELECT COUNT(*) FROM prezziario").fetchone()[0]
    if n_prez == 0:
        st.warning("Carica prima il prezziario."); st.stop()

    q = st.text_input("🔍 Cerca per codice o parola chiave",
                       placeholder="es: guaina  •  08.1  •  intonaco  •  lattoneria")
    c1,c2 = st.columns(2)
    cat_f = c1.selectbox("Categoria",["(tutte)"] + sorted(CATEGORY_MAP.values()))
    um_f  = c2.text_input("Filtra U.M.", placeholder="es: m2")

    with db() as conn:
        if q:
            rows=[dict(r) for r in conn.execute(
                "SELECT * FROM prezziario WHERE codice LIKE ? OR descrizione LIKE ? LIMIT 200",
                (f"%{q}%",f"%{q}%"))]
        else:
            rows=[dict(r) for r in conn.execute("SELECT * FROM prezziario LIMIT 200")]
    if cat_f != "(tutte)": rows=[r for r in rows if r.get("categoria")==cat_f]
    if um_f: rows=[r for r in rows if r.get("um","").lower()==um_f.lower()]

    # Fuzzy se pochi risultati
    if q and len(rows)<5:
        with db() as conn:
            all_items=[dict(r) for r in conn.execute("SELECT * FROM prezziario")]
        rows = fuzzy_search(all_items, q, limit=10)

    st.caption(f"Risultati: **{len(rows)}**")
    if not rows:
        st.info("Nessun risultato."); st.stop()

    df = pd.DataFrame(rows)[["id","codice","descrizione","um","prezzo","mo_pct","categoria"]]
    df.columns=["ID","Codice","Descrizione","U.M.","Prezzo €","MO %","Categoria"]
    st.dataframe(df, use_container_width=True, height=300,
                 column_config={
                     "Prezzo €":st.column_config.NumberColumn(format="€ %.2f"),
                     "MO %":st.column_config.NumberColumn(format="%.1f%%"),
                 })

    st.subheader("Aggiungi al computo")
    item_opts = {f"{r['codice']} – {r['descrizione'][:70]}": r for r in rows}
    sel_label = st.selectbox("Voce da aggiungere", list(item_opts.keys()))
    sel = item_opts[sel_label]

    with st.form("add_from_search"):
        c1,c2,c3 = st.columns(3)
        area = c1.text_input("Area", value="Zona A")
        cat_ = c2.selectbox("Categoria",["OG1","OG2","OG11","Sicurezza cantiere","Manodopera","Altro"])
        tipo = c3.selectbox("Tipo",["Soggetto a ribasso","Non soggetto a ribasso"])
        c4,c5,c6 = st.columns(3)
        qty  = c4.number_input("Quantità", min_value=0.0, value=1.0, format="%.3f")
        prz  = c5.number_input("Prezzo €", value=float(sel.get("prezzo",0)), format="%.2f")
        mo   = c6.number_input("MO %", value=float(sel.get("mo_pct",0)), min_value=0.0, max_value=100.0, format="%.1f")
        note = st.text_input("Note RUP")
        st.metric("Importo", f"€ {qty*prz:,.2f}")
        if st.form_submit_button("✅ Aggiungi al computo", type="primary"):
            imp=qty*prz; mo_e=imp*mo/100
            with db() as conn:
                conn.execute(
                    "INSERT INTO computo_voci(computo_id,area,categoria,tipo,codice,descrizione,um,qty,prezzo,importo,mo_pct,mo_euro,note) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (cid,area,cat_,tipo,sel["codice"],sel["descrizione"],sel["um"],qty,prz,imp,mo,mo_e,note))
                conn.commit()
            st.success(f"✅ Aggiunta: {sel['codice']}")


# ════════════════════════════════════════════════════════════════════════════
# 🧾 PREVENTIVO
# ════════════════════════════════════════════════════════════════════════════

elif page == "🧾 Preventivo":
    st.title("🧾 Carica Preventivo Ditta")
    cid = st.session_state.get("computo_id")
    if not cid: st.warning("Seleziona un computo."); st.stop()

    ditta = st.text_input("Nome ditta")
    up    = st.file_uploader("PDF del preventivo", type=["pdf"])
    if up and st.button("⚙️ Estrai voci"):
        pages,_ = extract_frames if False else (None,None)  # unused branch
        # usa pdftotext
        with tempfile.NamedTemporaryFile(suffix=".pdf",delete=False) as f:
            f.write(up.read()); tmp=f.name
        try:
            r=subprocess.run(["pdftotext","-layout",tmp,"-"],
                capture_output=True,text=True,encoding="utf-8",errors="replace")
            text=r.stdout
        except FileNotFoundError:
            text=""
        finally:
            os.unlink(tmp)
        lines=text.split("\n"); items=[]
        PREZZO_RE=re.compile(r'\b(\d{1,6}[.,]\d{2})\b')
        for line in lines:
            line=line.strip()
            if not line or len(line)<8: continue
            prices=PREZZO_RE.findall(line)
            if not prices: continue
            amounts=[pf(p) for p in prices]
            amount=amounts[-1]; up_=amounts[-2] if len(amounts)>=2 else amount
            fm=PREZZO_RE.search(line)
            desc=line[:fm.start()].strip() if fm else line
            if desc and len(desc)>5:
                items.append({"descrizione":desc[:200],"um":"","qty":None,
                              "prezzo":up_,"importo":amount,"mo_pct":0,"note":""})
        if items:
            st.success(f"Trovate {len(items)} voci.")
            st.session_state["prev_extracted"] = items
            st.session_state["prev_ditta"]     = ditta or up.name
        else:
            st.warning("Nessuna voce estratta. Inserisci manualmente.")

    if "prev_extracted" in st.session_state:
        st.subheader("Verifica voci estratte")
        df_e=pd.DataFrame(st.session_state["prev_extracted"])
        df_ed=st.data_editor(df_e,use_container_width=True,num_rows="dynamic",height=350,
            column_config={
                "descrizione":st.column_config.TextColumn("Descrizione",width="large"),
                "um":st.column_config.TextColumn("U.M.",width="small"),
                "qty":st.column_config.NumberColumn("Qty",format="%.3f"),
                "prezzo":st.column_config.NumberColumn("Prezzo €",format="%.2f"),
                "importo":st.column_config.NumberColumn("Importo €",format="%.2f"),
                "mo_pct":st.column_config.NumberColumn("MO %",format="%.1f"),
            })
        if st.button("💾 Salva preventivo", type="primary"):
            dname=st.session_state.get("prev_ditta","Ditta")
            with db() as conn:
                conn.execute("INSERT INTO preventivo(computo_id,ditta,filename) VALUES(?,?,?)",
                             (cid,dname,dname)); conn.commit()
                pid=conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                items_s=df_ed.to_dict(orient="records")
                items_s=[i for i in items_s if i.get("descrizione")]
                conn.executemany(
                    "INSERT INTO preventivo_voci(prev_id,descrizione,um,qty,prezzo,importo,mo_pct,note) VALUES(?,?,?,?,?,?,?,?)",
                    [(pid,i["descrizione"],i.get("um",""),i.get("qty"),
                      i.get("prezzo",0),i.get("importo",0),i.get("mo_pct",0),i.get("note","")) for i in items_s])
                conn.commit()
            del st.session_state["prev_extracted"]
            st.success(f"Salvato {len(items_s)} voci."); st.rerun()

    # inserimento manuale
    with db() as conn:
        prevs=[dict(r) for r in conn.execute("SELECT * FROM preventivo WHERE computo_id=?",(cid,))]
    if prevs:
        st.subheader("Preventivi salvati")
        for pr in prevs:
            with st.expander(f"🧾 {pr['ditta']} – {pr['creato'][:10]}"):
                with db() as conn:
                    pv=[dict(r) for r in conn.execute("SELECT * FROM preventivo_voci WHERE prev_id=?",(pr['id'],))]
                if pv:
                    st.dataframe(pd.DataFrame(pv)[["descrizione","um","qty","prezzo","importo"]],
                                 use_container_width=True, hide_index=True)
                    st.caption(f"Totale: € {sum(v['importo'] for v in pv if v['importo']):,.2f}")

    with st.expander("➕ Aggiungi voce manuale al preventivo"):
        if not prevs:
            st.info("Salva prima un preventivo.")
        else:
            p_opts={f"{p['ditta']}":p["id"] for p in prevs}
            p_sel=st.selectbox("Preventivo",list(p_opts.keys()))
            p_id=p_opts[p_sel]
            with st.form("add_prev_manual"):
                desc=st.text_area("Descrizione *")
                c1,c2,c3=st.columns(3)
                um_=c1.text_input("U.M."); qty_=c2.number_input("Qty",format="%.3f")
                prz_=c3.number_input("Prezzo €",format="%.2f")
                amt=st.number_input("Importo totale €",format="%.2f",value=qty_*prz_)
                if st.form_submit_button("Aggiungi"):
                    with db() as conn:
                        conn.execute("INSERT INTO preventivo_voci(prev_id,descrizione,um,qty,prezzo,importo) VALUES(?,?,?,?,?,?)",
                                     (p_id,desc,um_,qty_,prz_,amt)); conn.commit()
                    st.success("Aggiunta."); st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# ⚖️ CONFRONTO
# ════════════════════════════════════════════════════════════════════════════

elif page == "⚖️ Confronto":
    st.title("⚖️ Confronto Preventivo / Prezziario")
    st.caption("Il matching è un suggerimento. La scelta finale spetta al RUP.")
    cid=st.session_state.get("computo_id")
    if not cid: st.warning("Seleziona un computo."); st.stop()

    with db() as conn:
        prevs=[dict(r) for r in conn.execute("SELECT * FROM preventivo WHERE computo_id=?",(cid,))]
    if not prevs: st.warning("Carica prima un preventivo."); st.stop()

    p_opts={f"[{p['id']}] {p['ditta']}":p["id"] for p in prevs}
    p_sel=st.selectbox("Preventivo",list(p_opts.keys()))
    p_id=p_opts[p_sel]

    with db() as conn:
        pv=[dict(r) for r in conn.execute("SELECT * FROM preventivo_voci WHERE prev_id=?",(p_id,))]
        all_prez=[dict(r) for r in conn.execute("SELECT * FROM prezziario")]

    if not pv: st.info("Nessuna voce."); st.stop()

    top_n=st.slider("Suggerimenti per voce",3,10,5)
    tot_prev=0; tot_prez=0; da_chiarire=[]

    for vi in pv:
        q_pr=vi.get("prezzo",0) or 0
        qty=vi.get("qty") or 0
        amt=vi.get("importo",0) or (q_pr*qty)
        tot_prev+=amt
        matches=fuzzy_search(all_prez, vi.get("descrizione",""), limit=top_n)

        flags=[]
        if not vi.get("um"): flags.append("⚠️ U.M. assente")
        if not vi.get("qty"): flags.append("⚠️ Quantità assente")
        if q_pr==0: flags.append("⚠️ Prezzo assente")
        if len((vi.get("descrizione","") or "").split())<=3: flags.append("⚠️ Descrizione generica")

        label=f"{'⚠️ ' if flags else ''}**{vi['descrizione'][:80]}** | € {amt:,.2f}"
        with st.expander(label, expanded=False):
            if flags:
                for fl in flags: st.warning(fl)
            if not matches:
                st.info("Nessuna corrispondenza nel prezziario."); da_chiarire.append(vi); continue
            rows=[]
            for m in matches:
                pb_p=m.get("prezzo",0) or 0
                diff=q_pr-pb_p if pb_p else None
                dev=(diff/pb_p*100) if (diff is not None and pb_p) else None
                flag_=("🔴 +"+f"{dev:.0f}%" if dev and dev>20 else
                        "🟠 +"+f"{dev:.0f}%" if dev and dev>0 else
                        "🟢 "+f"{dev:.0f}%" if dev is not None else "⚠️ n/d")
                rows.append({"✓":False,"Codice":m.get("codice",""),
                             "Descrizione":m.get("descrizione","")[:100],
                             "U.M.":m.get("um",""),"Prezzo prezz.":pb_p,
                             "Prezzo prev.":q_pr,"Diff":diff,"Flag":flag_,
                             "_id":m.get("id"),"_pb_p":pb_p})
            df_m=pd.DataFrame(rows)
            edited=st.data_editor(df_m[["✓","Codice","Descrizione","U.M.","Prezzo prezz.","Prezzo prev.","Diff","Flag"]],
                hide_index=True, use_container_width=True,
                disabled=["Codice","Descrizione","U.M.","Prezzo prezz.","Prezzo prev.","Diff","Flag"],
                column_config={
                    "✓":st.column_config.CheckboxColumn("Sel.",width="small"),
                    "Prezzo prezz.":st.column_config.NumberColumn(format="€ %.2f"),
                    "Prezzo prev.":st.column_config.NumberColumn(format="€ %.2f"),
                    "Diff":st.column_config.NumberColumn(format="€ %.2f"),
                }, key=f"conf_{vi['id']}", height=min(40+38*len(rows),280))
            sel_rows=edited[edited["✓"]==True]
            if not sel_rows.empty:
                idx=sel_rows.index[0]
                pb_p2=df_m.iloc[idx]["_pb_p"]
                mid=df_m.iloc[idx]["_id"]
                dev2=(q_pr-pb_p2)/pb_p2*100 if pb_p2 else 0
                tot_prez+=pb_p2*qty if qty else pb_p2
                st.success(f"Selezionato: **{df_m.iloc[idx]['Codice']}**")
                if st.button("✅ Conferma abbinamento",key=f"save_{vi['id']}"):
                    with db() as conn:
                        conn.execute("UPDATE preventivo_voci SET match_id=?,scostamento=? WHERE id=?",
                                     (int(mid),round(dev2,2),vi["id"])); conn.commit()
                    st.success("Abbinamento salvato.")
            else:
                da_chiarire.append(vi)

    st.divider()
    c1,c2,c3=st.columns(3)
    c1.metric("Totale preventivo",f"€ {tot_prev:,.2f}")
    c2.metric("Totale prezziario",f"€ {tot_prez:,.2f}")
    diff_t=tot_prev-tot_prez
    c3.metric("Differenza",f"€ {diff_t:,.2f}",
              delta=f"{diff_t/tot_prez*100:.1f}%" if tot_prez else "n/d",
              delta_color="inverse")
    if da_chiarire:
        st.warning(f"**{len(da_chiarire)} voci** da chiarire / non abbinate:")
        for v in da_chiarire: st.write(f"• {v['descrizione'][:80]}")


# ════════════════════════════════════════════════════════════════════════════
# 🎬 SOPRALLUOGO VIDEO
# ════════════════════════════════════════════════════════════════════════════

elif page == "🎬 Sopralluogo Video":
    st.title("🎬 Sopralluogo Video")
    st.caption(
        "Carica un video di sopralluogo (ripresa del cantiere, del tetto, ecc.). "
        "L'app estrae un frame ogni N secondi: clicca su ogni frame per annotare gli interventi necessari. "
        "Gli interventi annotati possono essere importati direttamente nel computo."
    )
    cid=st.session_state.get("computo_id")
    if not cid: st.warning("Seleziona un computo."); st.stop()

    # Upload
    st.subheader("1. Carica il video")
    vid_file=st.file_uploader("Video di sopralluogo",type=["mp4","mov","avi","mkv","m4v"],
                               help="Formati supportati: MP4, MOV, AVI, MKV")
    col_int, col_max = st.columns(2)
    interval=col_int.slider("Estrai un frame ogni… (secondi)",1,30,5)
    max_fr  =col_max.slider("Numero massimo di frame",10,120,60)

    if vid_file and st.button("🎞️ Estrai frame", type="primary"):
        with st.spinner("Estrazione frame in corso…"):
            frames, msg = extract_frames(vid_file.read(), interval_sec=interval, max_frames=max_fr)
        if frames is None:
            st.error(msg)
            st.info("Per installare opencv: nel terminale digita `pip install opencv-python-headless`")
        else:
            st.success(msg)
            st.session_state["video_frames"]  = frames
            st.session_state["video_annots"]  = {}   # {frame_idx: {desc, priorita}}
            st.session_state["video_filename"]= vid_file.name

    # Galleria frame con annotazioni
    if "video_frames" in st.session_state and st.session_state["video_frames"]:
        frames  = st.session_state["video_frames"]
        annots  = st.session_state.get("video_annots", {})

        st.subheader("2. Annota gli interventi")
        st.info(
            "Per ogni frame che mostra un problema o un intervento necessario: "
            "espandi il pannello, scrivi la descrizione e la priorità."
        )

        n_annotated = sum(1 for a in annots.values() if a.get("desc"))
        st.metric("Frame annotati", f"{n_annotated} / {len(frames)}")

        # Mostra frame a 3 colonne
        cols_per_row = 3
        for row_start in range(0, len(frames), cols_per_row):
            cols = st.columns(cols_per_row)
            for ci, fi in enumerate(range(row_start, min(row_start+cols_per_row, len(frames)))):
                t_sec, jpg = frames[fi]
                mm, ss = divmod(int(t_sec), 60)
                has_ann = bool(annots.get(fi, {}).get("desc"))
                with cols[ci]:
                    st.image(jpg, caption=f"⏱ {mm:02d}:{ss:02d}" + (" ✅" if has_ann else ""),
                             use_container_width=True)
                    with st.expander("📝 Annota", expanded=has_ann):
                        prev_desc=annots.get(fi,{}).get("desc","")
                        prev_prio=annots.get(fi,{}).get("priorita","media")
                        desc_in=st.text_area("Intervento necessario",value=prev_desc,
                                             key=f"desc_{fi}", height=80,
                                             placeholder="es: Guaina ammalorata, rifacimento necessario")
                        prio_in=st.selectbox("Priorità",["alta","media","bassa"],
                                             index=["alta","media","bassa"].index(prev_prio),
                                             key=f"prio_{fi}")
                        if st.button("💾 Salva nota",key=f"save_ann_{fi}"):
                            if not isinstance(st.session_state.get("video_annots"),dict):
                                st.session_state["video_annots"]={}
                            st.session_state["video_annots"][fi]={
                                "desc":desc_in,"priorita":prio_in,"tempo":t_sec}
                            st.rerun()

        # Riepilogo annotazioni
        annotated = {fi:a for fi,a in annots.items() if a.get("desc")}
        if annotated:
            st.subheader("3. Riepilogo interventi annotati")
            rows_ann=[]
            for fi,a in sorted(annotated.items()):
                mm,ss=divmod(int(a["tempo"]),60)
                rows_ann.append({
                    "Frame":fi+1,"Tempo":f"{mm:02d}:{ss:02d}",
                    "Intervento":a["desc"],"Priorità":a["priorita"]
                })
            df_ann=pd.DataFrame(rows_ann)
            st.dataframe(df_ann,use_container_width=True,hide_index=True)

            st.subheader("4. Importa nel computo")
            st.caption("Seleziona gli interventi da importare come voci del computo.")
            with st.form("import_video"):
                area_v=st.text_input("Area intervento",value="Da sopralluogo video")
                cat_v =st.selectbox("Categoria",["OG1","OG2","OG11","Sicurezza cantiere","Altro"])
                tipo_v=st.selectbox("Tipo",["Soggetto a ribasso","Non soggetto a ribasso"])
                sels_v={}
                for fi,a in sorted(annotated.items()):
                    mm,ss=divmod(int(a["tempo"]),60)
                    sels_v[fi]=st.checkbox(
                        f"[{mm:02d}:{ss:02d}] {a['desc'][:80]} ({a['priorita']})",
                        value=True, key=f"import_chk_{fi}")
                if st.form_submit_button("✅ Importa voci selezionate nel computo", type="primary"):
                    count=0
                    with db() as conn:
                        for fi,chk in sels_v.items():
                            if chk:
                                a=annotated[fi]
                                mm2,ss2=divmod(int(a["tempo"]),60)
                                desc_c=f"{a['desc']} [video {mm2:02d}:{ss2:02d}]"
                                conn.execute(
                                    "INSERT INTO computo_voci(computo_id,area,categoria,tipo,codice,descrizione,um,qty,prezzo,importo,mo_pct,mo_euro,note) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                    (cid,area_v,cat_v,tipo_v,"",desc_c,"cad",1,0,0,0,0,
                                     f"Priorità: {a['priorita']}"))
                                count+=1
                        conn.commit()
                    st.success(f"✅ Importate {count} voci nel computo. Vai su **📋 Computo** per aggiungere prezzi e quantità.")
                    # pulisce
                    del st.session_state["video_frames"]
                    del st.session_state["video_annots"]
                    st.rerun()
        else:
            st.info("Nessuna annotazione ancora. Espandi i frame qui sopra per aggiungere note.")


# ════════════════════════════════════════════════════════════════════════════
# 📤 ESPORTA
# ════════════════════════════════════════════════════════════════════════════

elif page == "📤 Esporta":
    st.title("📤 Esporta Excel")
    cid=st.session_state.get("computo_id")
    if not cid: st.warning("Seleziona un computo."); st.stop()

    with db() as conn:
        meta=dict(conn.execute("SELECT * FROM computo_meta WHERE id=?",(cid,)).fetchone() or {})
        voci=[dict(r) for r in conn.execute("SELECT * FROM computo_voci WHERE computo_id=?",(cid,))]
        prevs=[dict(r) for r in conn.execute("SELECT * FROM preventivo WHERE computo_id=?",(cid,))]
        prez=[dict(r) for r in conn.execute("SELECT * FROM prezziario LIMIT 3000")]

    if not voci:
        st.warning("Il computo è vuoto. Aggiungi voci prima di esportare."); st.stop()

    tot=sum(v.get("importo",0) for v in voci)
    mo_=sum(v.get("mo_euro",0) for v in voci)
    rib=sum(v.get("importo",0) for v in voci if v.get("tipo")=="Soggetto a ribasso")
    nrib=sum(v.get("importo",0) for v in voci if v.get("tipo")=="Non soggetto a ribasso")

    c1,c2,c3,c4=st.columns(4)
    c1.metric("Totale computo",f"€ {tot:,.2f}")
    c2.metric("Manodopera",f"€ {mo_:,.2f}")
    c3.metric("A ribasso",f"€ {rib:,.2f}")
    c4.metric("Non a ribasso",f"€ {nrib:,.2f}")
    st.metric("Voci nel computo",len(voci))

    # Preventivo da includere
    prev_voci_exp=None
    if prevs:
        p_opts={"(nessuno)":None}
        p_opts.update({f"[{p['id']}] {p['ditta']}":p["id"] for p in prevs})
        p_sel=st.selectbox("Includi confronto preventivo",list(p_opts.keys()))
        p_id=p_opts[p_sel]
        if p_id:
            with db() as conn:
                pv=[dict(r) for r in conn.execute(
                    """SELECT pv.*, p2.codice as codice_match, p2.prezzo as prezzo_match,
                       (pv.prezzo - p2.prezzo) as diff
                       FROM preventivo_voci pv
                       LEFT JOIN prezziario p2 ON pv.match_id=p2.id
                       WHERE pv.prev_id=?""",(p_id,))]
            prev_voci_exp=pv

    if st.button("🔄 Genera Excel", type="primary"):
        with st.spinner("Generazione Excel…"):
            xlsx=export_excel(meta, voci,
                              prev_voci=prev_voci_exp,
                              prezziario=prez if len(prez)<2000 else None)
        fname=f"ComputoRUP_{(meta.get('titolo','computo') or 'computo').replace(' ','_')[:30]}_{datetime.today().strftime('%Y%m%d')}.xlsx"
        st.success("✅ Excel pronto!")
        st.download_button("⬇️ Scarica Excel",data=xlsx,file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # salva locale
        with open(os.path.join(EXPORT_DIR,fname),"wb") as f: f.write(xlsx)
        st.caption(f"Salvato anche in `exports/{fname}`")
